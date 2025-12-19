"""XLSX metadata extractor using openpyxl."""

from pathlib import Path
from typing import ClassVar

from openpyxl import load_workbook

from metaextract.core.models import DocumentMetadata, ExtractionResult, FileType
from metaextract.extractors.base import MetadataExtractor


class XLSXExtractor(MetadataExtractor):
    """Extract metadata from XLSX files using openpyxl."""

    SUPPORTED_TYPES: ClassVar[list[FileType]] = [FileType.XLSX]

    def extract(self) -> ExtractionResult:
        """Extract metadata from XLSX."""
        try:
            metadata = self._create_base_metadata()

            # Load workbook with data_only to avoid formula evaluation
            wb = load_workbook(str(self.file_path), read_only=True, data_only=True)
            props = wb.properties

            if props:
                # Extract creator
                if props.creator:
                    metadata.author = self._clean_string(props.creator)
                    if metadata.author and metadata.author not in metadata.users:
                        metadata.users.append(metadata.author)

                # Extract last modified by
                if props.lastModifiedBy:
                    metadata.last_modified_by = self._clean_string(props.lastModifiedBy)
                    if metadata.last_modified_by and metadata.last_modified_by not in metadata.users:
                        metadata.users.append(metadata.last_modified_by)

                # Extract dates
                metadata.created = props.created
                metadata.modified = props.modified

                # Store raw properties
                metadata.raw = {
                    "creator": str(props.creator) if props.creator else None,
                    "lastModifiedBy": str(props.lastModifiedBy) if props.lastModifiedBy else None,
                    "created": str(props.created) if props.created else None,
                    "modified": str(props.modified) if props.modified else None,
                    "title": str(props.title) if props.title else None,
                    "subject": str(props.subject) if props.subject else None,
                    "category": str(props.category) if props.category else None,
                    "revision": str(props.revision) if props.revision else None,
                }

            # Extract application info
            self._extract_app_info(metadata)

            wb.close()
            self._metadata = metadata
            return ExtractionResult(success=True, metadata=metadata)

        except Exception as e:
            return ExtractionResult(success=False, error=str(e))

    def _extract_app_info(self, metadata: DocumentMetadata) -> None:
        """Extract application info from the XLSX package."""
        import zipfile
        from lxml import etree

        try:
            with zipfile.ZipFile(self.file_path, "r") as zf:
                if "docProps/app.xml" in zf.namelist():
                    app_xml = zf.read("docProps/app.xml")
                    root = etree.fromstring(app_xml)

                    ns = {"ep": "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"}

                    app_elem = root.find("ep:Application", ns)
                    if app_elem is not None and app_elem.text:
                        metadata.application = app_elem.text
                        if app_elem.text not in metadata.software:
                            metadata.software.append(app_elem.text)

                    version_elem = root.find("ep:AppVersion", ns)
                    if version_elem is not None and version_elem.text:
                        metadata.app_version = version_elem.text

        except Exception:
            pass

    def extract_text(self) -> str | None:
        """Extract text content. Limited for spreadsheets."""
        try:
            wb = load_workbook(str(self.file_path), read_only=True, data_only=True)
            texts = []

            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            texts.append(str(cell))

            wb.close()
            return " ".join(texts)
        except Exception:
            return None
